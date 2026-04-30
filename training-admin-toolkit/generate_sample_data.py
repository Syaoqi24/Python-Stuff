"""
Generate sample data files for demonstrating training-admin-toolkit.
Run this once to create:
  - participants.xlsx  (with fake hyperlinks in columns K, L, M)
  - participants.csv   (names only, for split_pdf demo)
  - config_example.json

Usage:
    python generate_sample_data.py
"""

import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


FAKE_DRIVE_ID = "1BxiMVs0XRA5nFMdKvBdBZjgmUUqptlbs"

PARTICIPANTS = [
    ("1", "P001", "Budi Santoso",       "Teknik Komputer"),
    ("2", "P002", "Siti Rahayu",        "Administrasi"),
    ("3", "P003", "Ahmad Fauzi",        "Teknik Elektro"),
    ("4", "P004", "Dewi Kusumawati",    "Akuntansi"),
    ("5", "P005", "Rizky Pratama",      "Desain Grafis"),
]

HEADERS = [
    "NO", "ID", "NAMA LENGKAP", "PROGRAM",
    "COL_E", "COL_F", "COL_G", "COL_H", "COL_I", "COL_J",
    "FOTO", "FOTO IJAZAH", "LINK KTP"
]


def gdrive_fake_url(filename: str) -> str:
    return f"https://drive.google.com/file/d/{FAKE_DRIVE_ID}/view?usp=sharing"


def main():
    out_dir = Path(".")

    # ── participants.xlsx ──────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "DATA PESERTA"

    # Title row
    ws.merge_cells("A1:M1")
    ws["A1"] = "DAFTAR PESERTA PELATIHAN — SAMPLE DATA (DEMO)"
    ws["A1"].font      = Font(bold=True, size=12)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Header row
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor="2E75B6")
        cell.alignment = Alignment(horizontal="center")

    # Data rows with fake hyperlinks in columns K, L, M
    for row_idx, (no, pid, name, program) in enumerate(PARTICIPANTS, start=3):
        ws.cell(row=row_idx, column=1,  value=no)
        ws.cell(row=row_idx, column=2,  value=pid)
        ws.cell(row=row_idx, column=3,  value=name)
        ws.cell(row=row_idx, column=4,  value=program)

        # Columns K(11), L(12), M(13) — display text + fake hyperlink
        for col_idx, file_type in [(11, "FOTO"), (12, "IJAZAH"), (13, "KTP")]:
            filename   = f"{name.replace(' ', '_')}_{file_type}.jpg"
            cell       = ws.cell(row=row_idx, column=col_idx, value=filename)
            cell.hyperlink = gdrive_fake_url(filename)
            cell.font  = Font(color="0563C1", underline="single")

    # Column widths
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 18
    for col in ["K", "L", "M"]:
        ws.column_dimensions[col].width = 30

    xlsx_path = out_dir / "participants.xlsx"
    wb.save(xlsx_path)
    print(f"✓ Created: {xlsx_path}")

    # ── participants.csv ───────────────────────────────────────────────────────
    csv_path = out_dir / "participants.csv"
    with open(csv_path, "w", encoding="utf-8") as f:
        f.write("NO,NAMA LENGKAP\n")
        for no, _, name, _ in PARTICIPANTS:
            f.write(f"{no},{name}\n")
    print(f"✓ Created: {csv_path}")

    # ── config_example.json ────────────────────────────────────────────────────
    config = {
        "_comment": "Column indices are 1-based (column A = 1, B = 2, ...)",
        "name_column": 3,
        "header_row":  2,
        "data_start":  3,
        "columns": {
            "PHOTO":   11,
            "DIPLOMA": 12,
            "ID_CARD": 13
        }
    }
    cfg_path = out_dir / "config_example.json"
    with open(cfg_path, "w", encoding="utf-8") as f:
        json.dump(config, f, indent=4)
    print(f"✓ Created: {cfg_path}")

    print("\nSample data ready. Run:")
    print("  python download_from_spreadsheet.py --file participants.xlsx --skip-errors")
    print("  python split_pdf_by_participants.py --pdf your.pdf --names participants.csv --skip-header")


if __name__ == "__main__":
    main()
