"""
Download files from Google Drive hyperlinks embedded in an Excel spreadsheet.

Reads a .xlsx file, extracts Google Drive hyperlinks from specified columns,
and downloads each file into organized subfolders — one folder per column type.

Works with any spreadsheet where:
- One column contains participant/person names
- Other columns contain cells with Google Drive hyperlinks (photo, document, ID, etc.)

Usage:
    python download_from_spreadsheet.py
    python download_from_spreadsheet.py --file participants.xlsx
    python download_from_spreadsheet.py --file participants.xlsx --output ./downloads
    python download_from_spreadsheet.py --file participants.xlsx --config my_config.json
    python download_from_spreadsheet.py --file participants.xlsx --skip-errors

Config file (JSON) — optional, overrides defaults:
    {
        "name_column": 3,
        "header_row": 2,
        "data_start": 3,
        "columns": {
            "PHOTO":    11,
            "ID_CARD":  12,
            "DIPLOMA":  13
        }
    }
"""

import re
import sys
import json
import time
import argparse
import requests
from pathlib import Path
from openpyxl import load_workbook


# ── Default config (used when no --config file is provided) ───────────────────

DEFAULT_CONFIG = {
    "name_column": 3,       # Column index (1-based) containing person names
    "header_row":  2,       # Row containing column headers (skipped)
    "data_start":  3,       # First row containing actual data
    "columns": {            # {folder_name: column_index} for file downloads
        "PHOTO":    11,
        "ID_CARD":  12,
        "DIPLOMA":  13,
    }
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def load_config(config_path: str) -> dict:
    """Load column config from JSON file, fallback to defaults."""
    if config_path and Path(config_path).exists():
        with open(config_path, "r", encoding="utf-8") as f:
            cfg = json.load(f)
        print(f"[INFO] Config loaded from: {config_path}")
        return cfg
    if config_path:
        print(f"[WARN] Config file not found: {config_path} — using defaults")
    return DEFAULT_CONFIG


def sanitize(name: str) -> str:
    """Remove characters that are illegal in filenames."""
    return re.sub(r'[\\/:*?"<>|]', "_", str(name)).strip()


def unique_path(directory: Path, filename: str) -> Path:
    """Return a non-colliding path by appending (1), (2), etc. if needed."""
    target = directory / filename
    if not target.exists():
        return target
    stem, suffix = Path(filename).stem, Path(filename).suffix
    counter = 1
    while True:
        candidate = directory / f"{stem} ({counter}){suffix}"
        if not candidate.exists():
            return candidate
        counter += 1


def gdrive_direct_url(url: str) -> str:
    """Convert any Google Drive share URL to a direct download URL."""
    # Format: /file/d/{ID}/view  or  /file/d/{ID}/edit
    m = re.search(r"/file/d/([a-zA-Z0-9_-]+)", url)
    if m:
        return f"https://drive.google.com/uc?export=download&id={m.group(1)}"
    # Format: ?id={ID}
    m = re.search(r"[?&]id=([a-zA-Z0-9_-]+)", url)
    if m:
        return f"https://drive.google.com/uc?export=download&id={m.group(1)}"
    # Already a direct link or unrecognised format — return unchanged
    return url


def download_file(url: str, dest: Path, timeout: int = 60) -> int:
    """
    Download a file from URL to dest path.
    Handles Google Drive's virus-scan confirmation page for large files.
    Returns number of bytes written.
    """
    session = requests.Session()
    headers = {"User-Agent": "Mozilla/5.0 (compatible; gdrive-downloader/2.0)"}

    response = session.get(url, stream=True, timeout=timeout, headers=headers)
    response.raise_for_status()

    # Google Drive sometimes returns an HTML confirmation page for large files
    content_type = response.headers.get("Content-Type", "")
    if "text/html" in content_type:
        token_match = re.search(r'confirm=([0-9A-Za-z_-]+)', response.text)
        confirm_param = token_match.group(1) if token_match else "t"
        confirmed_url = url + f"&confirm={confirm_param}"
        response = session.get(confirmed_url, stream=True, timeout=timeout, headers=headers)
        response.raise_for_status()

    total_bytes = 0
    with open(dest, "wb") as f:
        for chunk in response.iter_content(chunk_size=8192):
            f.write(chunk)
            total_bytes += len(chunk)
    return total_bytes


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description=(
            "Download files from Google Drive hyperlinks embedded in an Excel spreadsheet.\n"
            "Each column type is saved into its own subfolder under the output directory."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--file", "-f",
        default="participants.xlsx",
        help="Path to the .xlsx file  (default: participants.xlsx)"
    )
    parser.add_argument(
        "--output", "-o",
        default="./downloads",
        help="Root output folder  (default: ./downloads)"
    )
    parser.add_argument(
        "--config", "-c",
        default=None,
        help="Path to JSON config file specifying column layout  (optional)"
    )
    parser.add_argument(
        "--delay",
        type=float,
        default=1.0,
        help="Seconds to wait between downloads to avoid rate-limiting  (default: 1.0)"
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=60,
        help="HTTP timeout per download in seconds  (default: 60)"
    )
    parser.add_argument(
        "--skip-errors",
        action="store_true",
        help="Continue processing even if some downloads fail"
    )
    args = parser.parse_args()

    # ── Load config ───────────────────────────────────────────────────────────
    cfg         = load_config(args.config)
    NAME_COLUMN = cfg.get("name_column", DEFAULT_CONFIG["name_column"])
    DATA_START  = cfg.get("data_start",  DEFAULT_CONFIG["data_start"])
    COLUMN_MAP  = cfg.get("columns",     DEFAULT_CONFIG["columns"])

    # ── Load workbook ─────────────────────────────────────────────────────────
    fp = Path(args.file)
    if not fp.exists():
        sys.exit(
            f"\n[ERROR] File not found: {fp}\n"
            f"        Make sure '{fp.name}' is in the same folder as this script,\n"
            f"        or pass the full path with --file.\n"
        )

    print(f"\n[INFO] Loading : {fp.resolve()}")
    wb = load_workbook(fp)
    ws = wb.active
    print(f"[INFO] Sheet   : '{ws.title}'  ({ws.max_row} rows, {ws.max_column} columns)")
    print(f"[INFO] Columns : {COLUMN_MAP}")
    print()

    # ── Create output folders ─────────────────────────────────────────────────
    root    = Path(args.output)
    folders = {}
    for folder_name in COLUMN_MAP:
        folder = root / folder_name
        folder.mkdir(parents=True, exist_ok=True)
        folders[folder_name] = folder

    print("Output folders created:")
    for name, folder in folders.items():
        print(f"  {folder.resolve()}")
    print()

    # ── Process rows ──────────────────────────────────────────────────────────
    totals = {name: {"ok": 0, "skip": 0, "fail": 0} for name in COLUMN_MAP}

    for row_num in range(DATA_START, ws.max_row + 1):
        name_cell   = ws.cell(row=row_num, column=NAME_COLUMN)
        person_name = str(name_cell.value or "").strip()

        if not person_name or person_name.lower() in ("none", ""):
            continue

        for folder_name, col_idx in COLUMN_MAP.items():
            cell    = ws.cell(row=row_num, column=col_idx)
            display = str(cell.value or "").strip()

            # Extract hyperlink URL from the cell
            raw_url = cell.hyperlink.target if cell.hyperlink else None

            if not raw_url:
                if display:
                    print(f"  [{folder_name}] {person_name}: no hyperlink on '{display}' — skipped")
                totals[folder_name]["skip"] += 1
                continue

            dl_url   = gdrive_direct_url(raw_url)
            filename = sanitize(display) if display else f"row{row_num}"
            dest     = unique_path(folders[folder_name], filename)

            print(f"  [{folder_name}] {person_name}")
            print(f"    File  : {display}")
            print(f"    → {dest.relative_to(root)}", end=" ", flush=True)

            try:
                size_kb = download_file(dl_url, dest, timeout=args.timeout) / 1024
                print(f"({size_kb:.1f} KB) ✓")
                totals[folder_name]["ok"] += 1
            except Exception as exc:
                print(f"✗\n    [FAIL] {exc}")
                totals[folder_name]["fail"] += 1
                if not args.skip_errors:
                    sys.exit("\n[STOPPED] Use --skip-errors to continue past failures.\n")

            time.sleep(args.delay)

    # ── Summary ───────────────────────────────────────────────────────────────
    print(f"\n{'═' * 58}")
    print("  DOWNLOAD SUMMARY")
    print(f"{'─' * 58}")
    for name in COLUMN_MAP:
        t = totals[name]
        print(f"  {name:<14}  ✓ {t['ok']:>3} ok   ✗ {t['fail']:>2} failed   ○ {t['skip']:>2} skipped")
    print(f"{'─' * 58}")
    total_ok   = sum(t["ok"]   for t in totals.values())
    total_fail = sum(t["fail"] for t in totals.values())
    print(f"  Total downloaded : {total_ok}   Failed : {total_fail}")
    print(f"  Saved to         : {root.resolve()}")
    print(f"{'═' * 58}\n")

    print("Folder structure:")
    for name, folder in folders.items():
        count = sum(1 for _ in folder.iterdir())
        print(f"  {root}/{name}/   ({count} file(s))")
    print()


if __name__ == "__main__":
    main()
